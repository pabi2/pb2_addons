# -*- coding: utf-8 -*-
import base64
import cStringIO
from datetime import datetime

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection
from openpyxl.utils import quote_sheetname

from openerp import models, fields, api, _
from openerp.exceptions import Warning as UserError

SHEET_FORMULAS = {}


class ExportAsseItem(models.TransientModel):
    _name = 'export.asset.item'

    @api.model
    def _default_plan_template(self):
        template = self.env['ir.attachment'].\
            search([('budget_template', '=', True)])
        return template

    attachment_id = fields.Many2one(
        'ir.attachment',
        string='Template Plan',
        default=_default_plan_template,
    )

    @api.model
    def _update_program_group_sheet(self, workbook):
        try:
            PGObj = self.env['res.program.group']
            PG_Sheet = workbook.get_sheet_by_name('master_program_group')
            PG_Sheet.protection.sheet = True
            PG_Sheet.protection.set_password('pabi2')
            pg_ids = PGObj.search([])
            # create lines in activity group data sheet
            row = 2
            for pg in pg_ids:
                PG_Sheet.cell(row=row, column=1, value=pg.name)
                row += 1
            formula1 = "{0}!$A$2:$A$%s" % (row)
            PGList = DataValidation(
                type="list",
                formula1=formula1.format(
                    quote_sheetname('master_program_group')
                )
            )
            SHEET_FORMULAS.update({'formula_program_group': PGList})
        except:
            pass

    @api.model
    def _update_invest_asset_category_sheet(self, workbook):
        try:
            IACObj = self.env['res.invest.asset.category']
            IAC_Sheet = workbook.get_sheet_by_name('master_asset_category')
            IAC_Sheet.protection.sheet = True
            IAC_Sheet.protection.set_password('pabi2')
            iac_ids = IACObj.search([])
            # create lines in activity group data sheet
            row = 2
            for iac in iac_ids:
                IAC_Sheet.cell(row=row, column=1, value=iac.name)
                row += 1
            formula1 = "{0}!$A$2:$A$%s" % (row)
            IACList = DataValidation(
                type="list",
                formula1=formula1.format(
                    quote_sheetname('master_asset_category')
                )
            )
            SHEET_FORMULAS.update({'formula_asset_category': IACList})
        except:
            pass

    @api.model
    def _update_section_sheet(self, workbook):
        try:
            SectionObj = self.env['res.section']
            Section_Sheet = workbook.get_sheet_by_name('master_section')
            Section_Sheet.protection.sheet = True
            Section_Sheet.protection.set_password('pabi2')
            section_ids = SectionObj.search([])
            # create lines in activity group data sheet
            row = 2
            for section in section_ids:
                Section_Sheet.cell(row=row, column=1, value=section.name)
                row += 1
            formula1 = "{0}!$A$2:$A$%s" % (row)
            SectionList = DataValidation(
                type="list",
                formula1=formula1.format(
                    quote_sheetname('master_section')
                )
            )
            SHEET_FORMULAS.update({'formula_section': SectionList})
        except:
            pass

    @api.model
    def _update_division_sheet(self, workbook):
        try:
            DivisionObj = self.env['res.division']
            Division_Sheet = workbook.get_sheet_by_name('master_division')
            Division_Sheet.protection.sheet = True
            Division_Sheet.protection.set_password('pabi2')
            division_ids = DivisionObj.search([])
            # create lines in activity group data sheet
            row = 2
            for division in division_ids:
                Division_Sheet.cell(row=row, column=1, value=division.name)
                row += 1
            formula1 = "{0}!$A$2:$A$%s" % (row)
            DivisionList = DataValidation(
                type="list",
                formula1=formula1.format(
                    quote_sheetname('master_division')
                )
            )
            SHEET_FORMULAS.update({'formula_division': DivisionList})
        except:
            pass

    @api.model
    def _update_employees_sheet(self, workbook):
        try:
            EmployeeObj = self.env['hr.employee']
            Employee_Sheet = workbook.get_sheet_by_name('master_employee')
            Employee_Sheet.protection.sheet = True
            Employee_Sheet.protection.set_password('pabi2')
            employee_ids = EmployeeObj.search([])

            row = 2
            for employee in employee_ids:
                Employee_Sheet.cell(row=row, column=1, value=employee.name)
                row += 1
            formula1 = "{0}!$A$2:$A$%s" % (row)
            EmployeeList = DataValidation(
                type="list",
                formula1=formula1.format(
                    quote_sheetname('master_employee')
                )
            )
            SHEET_FORMULAS.update({'formula_employee': EmployeeList})
        except:
            pass

    @api.model
    def _format_sheet(self, Sheet):
        Sheet.protection.sheet = True
        Sheet.protection.set_password('pabi2')

        first_row = 8
        last_row = first_row + 100
        Reason = DataValidation(
            type="list",
            formula1='"New,Replacement,Extra"'
        )
        SHEET_FORMULAS.update({'reason': Reason})
        Sheet.add_data_validation(Reason)

        ResProgramGroup = SHEET_FORMULAS.get('formula_program_group', False)
        Sheet.add_data_validation(ResProgramGroup)
        ResInvestAsset = SHEET_FORMULAS.get('formula_asset_category', False)
        Sheet.add_data_validation(ResInvestAsset)
        ResSection = SHEET_FORMULAS.get('formula_section', False)
        Sheet.add_data_validation(ResSection)
        ResDivision = SHEET_FORMULAS.get('formula_division', False)
        Sheet.add_data_validation(ResDivision)
        Employee = SHEET_FORMULAS.get('formula_employee', False)
        Sheet.add_data_validation(Employee)

        protection = Protection(locked=False)

        for row in range(first_row, last_row):
            for col in range(1, 30):
                if col == 2:
                    ResProgramGroup.add(Sheet.cell(row=row, column=col))
                if col == 3:
                    ResInvestAsset.add(Sheet.cell(row=row, column=col))
                    Sheet.cell(row=row, column=col).protection = protection
                if col == 6:
                    Employee.add(Sheet.cell(row=row, column=col))
                if col == 7:
                    ResSection.add(Sheet.cell(row=row, column=col))
                if col == 8:
                    ResDivision.add(Sheet.cell(row=row, column=col))

                if col == 15:
                    Reason.add(Sheet.cell(row=row, column=col))

                if col == 12:
                    # Price subtotal
                    value = "=J%s*$K$%s" % (row, row)
                    Sheet.cell(row=row, column=col).value = value
                elif col == 14:
                    # Price Total
                    value = "=L%s+$M$%s" % (row, row)
                    Sheet.cell(row=row, column=col).value = value
                elif col == 22:
                    # Total Commitment
                    value = "=W%s+$X$%s+$Y$%s" % (row, row, row)
                    Sheet.cell(row=row, column=col).value = value
                elif col == 27:
                    # Total Commitment + Actual
                    value = "=V%s+$Z$%s" % (row, row)
                    Sheet.cell(row=row, column=col).value = value
                elif col == 29:
                    # Carried Forward (Total Commitment + Residual Budget)
                    value = "=V%s+$AB$%s" % (row, row)
                    Sheet.cell(row=row, column=col).value = value
                else:
                    Sheet.cell(row=row, column=col).protection = protection

    @api.model
    def _update_sheet_lines(self, Sheet, lines):
        row = 8
        for line in lines:
            if line.program_group_id:
                Sheet.cell(row=row, column=2, value=line.program_group_id.name)
            if line.invest_asset_categ_id:
                Sheet.cell(row=row, column=3,
                           value=line.invest_asset_categ_id.name)
            if line.asset_common_name:
                Sheet.cell(row=row, column=4, value=line.asset_common_name)
            if line.asset_name:
                Sheet.cell(row=row, column=5, value=line.asset_name)
            if line.request_user_id:
                Sheet.cell(row=row, column=6, value=line.request_user_id.name)
            if line.section_id:
                Sheet.cell(row=row, column=7, value=line.section_id.name)
            if line.division_id:
                Sheet.cell(row=row, column=8, value=line.division_id.name)
            if line.location:
                Sheet.cell(row=row, column=9, value=line.location)
            Sheet.cell(row=row, column=10, value=line.quantity)
            Sheet.cell(row=row, column=11, value=line.price_unit)
#             Sheet.cell(row=row, column=12, value='')
            Sheet.cell(row=row, column=13, value=line.price_other)
#             Sheet.cell(row=row, column=14, value='')

            if line.reason_purchase == 'new':
                Sheet.cell(row=row, column=15, value='New')
            elif line.reason_purchase == 'replace':
                Sheet.cell(row=row, column=15, value='Replacement')
            elif line.reason_purchase == 'extra':
                Sheet.cell(row=row, column=15, value='Extra')

            if line.reason_purchase_text:
                Sheet.cell(row=row, column=16, value=line.reason_purchase_text)
#             Sheet.cell(row=row, column=17, value='')
            Sheet.cell(row=row, column=18, value=line.planned_utilization)
            if line.quotation_document:
                Sheet.cell(row=row, column=19, value=line.quotation_document)
            if line.specification_summary:
                Sheet.cell(row=row, column=20,
                           value=line.specification_summary)
#             Sheet.cell(row=row, column=21, value='')
#             Sheet.cell(row=row, column=22, value='')
            Sheet.cell(row=row, column=23, value=line.po_commitment)
            Sheet.cell(row=row, column=24, value=line.po_commitment)
            Sheet.cell(row=row, column=25, value=line.exp_commitment)
            Sheet.cell(row=row, column=26, value=line.actual_amount)
#             Sheet.cell(row=row, column=27, value='')
            Sheet.cell(row=row, column=28, value=line.budget_residual)
            row += 1
        return True

    @api.model
    def _update_header(self, Sheet, item):
        Sheet.cell(row=1, column=5, value=item.id)
        org = item.org_id.code or item.org_id.name_short
        Sheet.cell(row=1, column=2, value=item.fiscalyear_id.name)
        Sheet.cell(row=2, column=2, value=org)
        Sheet.cell(row=3, column=2,
                   value=datetime.today().strftime('%d-%m-%Y'))
        Sheet.cell(row=4, column=2, value=self.env.user.name)

    @api.model
    def _get_filename(self, item):
        template_file = self.attachment_id
        org = item.org_id.code or item.org_id.name_short
        filename = '%s-%s-%s-%s.xlsx' % ('AssetItem',
                                         item.fiscalyear_id.name,
                                         org,
                                         template_file.name)
        return filename

    @api.multi
    def action_export(self):
        active_ids = self.env.context.get('active_ids', [])
        active_model = self.env.context.get('active_model', False)
        items = self.env[active_model].browse(active_ids)

        if not self.attachment_id:
            raise UserError(_('Please add .xlsx template.'))
        if not items:
            raise UserError(_('No items to export.'))

        for item in items:
            export_file = self.attachment_id.datas.decode('base64')
            stream = cStringIO.StringIO(export_file)
            workbook = openpyxl.load_workbook(stream)

            self._update_program_group_sheet(workbook)
            self._update_invest_asset_category_sheet(workbook)
            self._update_section_sheet(workbook)
            self._update_division_sheet(workbook)
            self._update_employees_sheet(workbook)

            Asset_Sheet = workbook.get_sheet_by_name('Assets')
            self._update_header(Asset_Sheet, item)
            self._format_sheet(Asset_Sheet)
            self._update_sheet_lines(Asset_Sheet, item.item_ids)

            filename = self._get_filename(item)
            stream1 = cStringIO.StringIO()
            workbook.save(stream1)

            attachement_id = self.env['ir.attachment'].create({
                'name': filename,
                'datas': stream1.getvalue().encode('base64'),
                'datas_fname': filename,
                'res_model': 'invest.asset.plan',
                'res_id': item.id,
                'invest_asset_plan_id': item.id,
                'description': 'Export',
            })
            output_file = self.env['budget.xls.output'].create({
                'name': filename,
                'xls_output': base64.encodestring(stream1.getvalue()),
            })
            self.env['budget.plan.history'].create({
                'user_id': self.env.user.id,
                'operation_date': fields.Datetime.now(),
                'operation_type': 'export',
                'invest_asset_plan_id': item.id,
                'attachement_id': attachement_id.id
            })

        return {
            'context': self.env.context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'budget.xls.output',
            'res_id': output_file.id,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }
