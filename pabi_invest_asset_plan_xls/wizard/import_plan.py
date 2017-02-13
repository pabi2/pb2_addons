# -*- coding: utf-8 -*-
# © <YEAR(S)> <AUTHOR(S)>
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl.html).
import sys
import cStringIO
from datetime import datetime

import openpyxl

from openerp import tools
from openerp import models, fields, api, _
from openerp.exceptions import Warning as UserError

class ImportAsseItem(models.TransientModel):
    _name = 'import.asset.item'

    input_file = fields.Binary('Template')
    datas_fname = fields.Char('Import File Name')
    
    
    @api.multi
    def import_budget(self):
        active_ids = self.env.context.get('active_ids', [])
        active_model = self.env.context.get('active_model', False)
        
        imp_file = self.input_file.decode('base64')
        stream = cStringIO.StringIO(imp_file)

        try:
            workbook = openpyxl.load_workbook(stream)
        except IOError as e:
            raise UserError(_(e.strerror))
        except ValueError as e:
            raise UserError(_(e.strerror))
        except:
            e = sys.exc_info()[0]
            raise UserError(_('Wrong file format. Please enter .xlsx file.'))
        Plans = self.env[active_model].browse(active_ids)
        for plan in Plans:
            if plan.state != 'draft':
                raise UserError(
                    _('You can update budget plan only in draft state!'))

            Asset_Sheet = workbook.get_sheet_by_name('Assets')
            vals = {}
            plan_id = Asset_Sheet.cell(row=1, column=5).value
            # if we trying to import sheet of other record then raise error
            if plan.id != plan_id:
                raise UserError(
                    _('Please import the correct file for this plan')
                )
                    # get fiscal year from sheet
            fiscal_year = Asset_Sheet.cell(row=1, column=2).value
            fiscal_year_id = self.env['account.fiscalyear'].search(
                [('name', '=', tools.ustr(fiscal_year))])
            if fiscal_year_id:
                vals.update({'fiscalyear_id': fiscal_year_id.id})
            # get org from sheet
            org = Asset_Sheet.cell(row=2, column=2).value
            org_id =\
                self.env['res.org'].search(
                    ['|',
                     ('code', '=', tools.ustr(org)),
                     ('name_short', '=', tools.ustr(org))])
            export_date = Asset_Sheet.cell(row=3, column=2).value
            if export_date:
                export_date = datetime.strftime(
                                datetime.strptime(
                                    export_date, '%d-%m-%Y'), '%Y-%m-%d')
                vals.update({'date': export_date})
            responsible_by = Asset_Sheet.cell(row=5, column=2).value
            responsible_by_id = self.env['res.users'].search(
                [('name', '=', tools.ustr(responsible_by))])
            if responsible_by_id:
                vals.update({'creating_user_id': responsible_by_id.id})

            line_row = 8
            lines_to_create = []

            max_row = Asset_Sheet.max_row
            for row in range(line_row, max_row):
                line_vals = {}
                program_group = Asset_Sheet.cell(row=row, column=2).value
                program_group_id = self.env['res.program.group'].search(
                    [('name', '=', tools.ustr(program_group))])
                if program_group_id:
                    line_vals.update({'program_group_id': program_group_id.id})

                asset_category = Asset_Sheet.cell(row=row, column=3).value
                asset_category_id = self.env['res.invest.asset.category'].search(
                    [('name', '=', tools.ustr(asset_category))])
                if asset_category_id:
                    line_vals.update({'invest_asset_categ_id': asset_category_id.id})

                if not program_group_id and not asset_category_id:
                    continue

                asset_common_name = Asset_Sheet.cell(row=row, column=4).value
                if asset_common_name == '=FALSE()':
                    asset_common_name = ''
                line_vals.update({'asset_common_name': asset_common_name})

                asset_name = Asset_Sheet.cell(row=row, column=5).value
                if asset_name == '=FALSE()':
                    asset_name = ''
                line_vals.update({'asset_name': asset_name})

                requester = Asset_Sheet.cell(row=row, column=6).value
                requester = tools.ustr(requester)
                requester_code = requester[1:7]
                requester_id = self.env['hr.employee'].search(
                    [('employee_code', '=', requester_code)])
                if requester_id:
                    line_vals.update({'request_user_id': requester_id.id})

                section = Asset_Sheet.cell(row=row, column=7).value
                section_id = self.env['res.section'].search(
                    [('name', '=', tools.ustr(section))])
                if section_id:
                    line_vals.update({'section_id': section_id.id})

                division = Asset_Sheet.cell(row=row, column=8).value
                division_id = self.env['res.division'].search(
                    [('name', '=', tools.ustr(division))])
                if section_id:
                    line_vals.update({'division_id': division_id.id})

                location = Asset_Sheet.cell(row=row, column=9).value
                if location == '=FALSE()':
                    location = ''
                line_vals.update({'location': location})

                quantity = Asset_Sheet.cell(row=row, column=10).value or 0.0
                line_vals.update({'quantity': quantity})

                price_unit = Asset_Sheet.cell(row=row, column=11).value or 0.0
                line_vals.update({'price_unit': price_unit})

                price_other = Asset_Sheet.cell(row=row, column=13).value or 0.0
                line_vals.update({'price_other': price_other})

                reason_purchase = Asset_Sheet.cell(row=row, column=15).value
                if reason_purchase:
                    if tools.ustr(reason_purchase) == 'Replacement':
                        line_vals.update({'reason_purchase': 'replace'})
                    elif tools.ustr(reason_purchase) == 'Extra':
                        line_vals.update({'reason_purchase': 'extra'})
                    else:
                        line_vals.update({'reason_purchase': 'new'})

                reason_purchase_text = Asset_Sheet.cell(row=row, column=16).value
                if reason_purchase_text == '=FALSE()':
                    reason_purchase_text = ''
                line_vals.update({'reason_purchase_text': reason_purchase_text})

                planned_utilization = Asset_Sheet.cell(row=row, column=18).value
                if planned_utilization == '=FALSE()':
                    planned_utilization = ''
                line_vals.update({'planned_utilization': planned_utilization})

                quotation_document = Asset_Sheet.cell(row=row, column=19).value
                if quotation_document == '=FALSE()':
                    quotation_document = ''
                line_vals.update({'quotation_document': quotation_document})

                specification_summary = Asset_Sheet.cell(row=row, column=20).value
                if specification_summary == '=FALSE()':
                    specification_summary = ''
                line_vals.update({'specification_summary': specification_summary})

                pr_commitment = Asset_Sheet.cell(row=row, column=23).value or 0.0
                line_vals.update({'pr_commitment': pr_commitment})

                po_commitment = Asset_Sheet.cell(row=row, column=24).value or 0.0
                line_vals.update({'po_commitment': po_commitment})

                exp_commitment = Asset_Sheet.cell(row=row, column=25).value or 0.0
                line_vals.update({'exp_commitment': exp_commitment})

                actual_amount = Asset_Sheet.cell(row=row, column=26).value or 0.0
                line_vals.update({'actual_amount': actual_amount})

                budget_residual = Asset_Sheet.cell(row=row, column=28).value or 0.0
                line_vals.update({'budget_residual': budget_residual})

                lines_to_create.append((0, 0, line_vals))

            vals.update({'item_ids': lines_to_create})
            existing_lines = self.env['invest.asset.plan.item'].\
                search([('plan_id', '=', plan.id)])
            if existing_lines:
                existing_lines.unlink()
            plan.write(vals)
            
            attachement_id = self.env['ir.attachment'].create({
                'name': self.datas_fname,
                'datas': stream.getvalue().encode('base64'),
                'datas_fname': self.datas_fname,
                'res_model': 'invest.asset.plan',
                'res_id': plan.id,
                'invest_asset_plan_id': plan.id,
                'description': 'Import',
            })
            
            
        return True