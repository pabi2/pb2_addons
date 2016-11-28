# -*- coding: utf-8 -*-
# © <YEAR(S)> <AUTHOR(S)>
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl.html).
import sys
import cStringIO

import openpyxl

from openerp import tools
from openerp import models, fields, api, _
from openerp.exceptions import Warning as UserError


class BudgetImportWizard(models.TransientModel):
    _name = 'budget.import.wizard'

    input_file = fields.Binary('Template')
    datas_fname = fields.Char('Import File Name')

    @api.model
    def _update_cost_control_lines(self, workbook, budget):
        CostCtrl_Sheet = workbook.get_sheet_by_name('JobOrder')
        max_row = CostCtrl_Sheet.max_row

        cc_row = 8
        for row in range(cc_row, max_row):
            costcontrol_vals = {'plan_id': budget.id}
            costcontrol = CostCtrl_Sheet.cell(row=cc_row, column=2).value
            if costcontrol:
                costcontrol_id =  self.env['cost.control'].search([('name', '=', tools.ustr(costcontrol))])
                if costcontrol_id:
                    costcontrol_vals.update({'cost_control_id': costcontrol_id.id})
                else:
                    raise UserError(_('Please select valid Job Order!'))
            else:
                cc_row += 28
                continue

            existing_costcontrol_lines =\
                self.env['budget.plan.unit.cost.control'].search(
                    [('plan_id','=' ,budget.id),
                     ('cost_control_id', '=', costcontrol_id.id)])
            if existing_costcontrol_lines:
                existing_costcontrol_lines.unlink()
            costcontrol_line = self.env['budget.plan.unit.cost.control'].create(costcontrol_vals)

            line_start = cc_row + 5
            calculate_vals = {}
            ids = []
            for row in range(line_start, line_start+20):
                line_vals = {'cost_control_line_id': costcontrol_line.id}
                col = 1
                activity_group = CostCtrl_Sheet.cell(row=row, column=col).value
                if activity_group:
                    activity_group_id =  self.env['account.activity.group'].search([('name', '=', tools.ustr(activity_group))])
                    line_vals.update({'activity_group_id': activity_group_id.id})
                else:
                    continue
                col += 1

                name = CostCtrl_Sheet.cell(row=row, column=col).value
                if name:
                    line_vals.update({'name': tools.ustr(name)})
                    calculate_vals.update({'description': tools.ustr(name)})
                col += 1
                col += 1

                unit = CostCtrl_Sheet.cell(row=row, column=col).value
                if unit and not isinstance(unit, long):
                    raise UserError(
                        _('Please insert float value on\
                         row: %s - column: %s') % (row, 5))
                if not unit:
                    unit = 0.0
                calculate_vals.update({'unit': unit})
                line_vals.update({'unit': unit})
                col += 1

                activity_unit_price = CostCtrl_Sheet.cell(row=row, column=col).value
                if activity_unit_price:
                    if activity_unit_price and not isinstance(activity_unit_price, long):
                        raise UserError(
                            _('Please insert float value on\
                             row: %s - column: %s') % (row, col))
                else:
                    activity_unit_price = 0.0
                calculate_vals.update({'activity_unit_price': activity_unit_price})
                line_vals.update({'activity_unit_price': activity_unit_price})
                col += 1
                activity_unit = CostCtrl_Sheet.cell(row=row, column=col).value
                if activity_unit:
                    if activity_unit and not isinstance(activity_unit, long):
                        raise UserError(
                            _('Please insert float value on\
                             row: %s - column: %s') % (row, col))
                else:
                    activity_unit = 0.0
                calculate_vals.update({'activity_unit': activity_unit})
                line_vals.update({'activity_unit': activity_unit})
                total_act_budget = unit * activity_unit_price * activity_unit
                col += 1
                col += 1
                col += 1
                m_col = col
                cnt = 1
                total_month_budget = 0
                for c in range(m_col, m_col+12):
                    val = CostCtrl_Sheet.cell(row=row, column=c).value
                    m_field = 'm'+str(cnt)
                    if val:
                        if val and not isinstance(val, long):
                            raise UserError(
                                _('Please insert float value on\
                                 row: %s - column: %s') % (row, col))
                        total_month_budget += val
                        line_vals.update({m_field : val})
                    cnt += 1
                    col += 1
                col += 1
                col += 1
                if (total_act_budget - total_month_budget) != 0.0:
                    raise UserError(_('Please verify budget lines total!'))
                cc_line = self.env['budget.plan.unit.cost.control.line'].create(line_vals)
                if cc_line:
                    plan_line_vals = {}
                    plan_line_vals.update(line_vals)
                    plan_line_vals.update(calculate_vals)
                    create_vals = {
                        'breakdown_line_id': cc_line.id,
                        'plan_id': costcontrol_line.plan_id.id,
                        'fk_costcontrol_id': costcontrol_line.id,
                        'section_id': costcontrol_line.plan_id.section_id.id,
                        'cost_control_id': costcontrol_line.cost_control_id.id,
                    }
                    create_vals.update(plan_line_vals)
                    new_id = self.env['budget.plan.unit.line'].create(create_vals).id
                    ids.append(new_id)
            cc_row += 28
            costcontrol_line.write({'detail_ids': [(6, 0, ids)]})
        return True

    @api.multi
    def update_budget_prepare(self, budget_ids, template=None):
        if not template:
            raise UserError(_('Please add .xlsx template.'))
        if not budget_ids:
            raise UserError(_('No budget to import.'))

        imp_file = template.decode('base64')
        stream = cStringIO.StringIO(imp_file)

        try:
            workbook = openpyxl.load_workbook(stream)
        except IOError as e:
            raise UserError(_(e.strerror))
        except ValueError as e:
            raise UserError(_(e.strerror))
        except:
            e = sys.exc_info()[0]
            raise UserError(_('Wrong file format. Please enter .xlsx file.'))
        budgets = self.env['budget.plan.unit'].browse(budget_ids)
        for budget in budgets:

            if budget.state != 'draft':
                raise UserError(
                    _('You can update budget plan only in draft state!'))

            NonCostCtrl_Sheet = workbook.get_sheet_by_name('Non_jobOrder')
            max_row = NonCostCtrl_Sheet.max_row
            vals = {}

            bg_id = NonCostCtrl_Sheet.cell(row=1, column=5).value
            if budget.id != bg_id:
                raise UserError(
                    _('Please import the correct file for this plan')
                )
            fiscal_year = NonCostCtrl_Sheet.cell(row=1, column=2).value
            fiscal_year_id = self.env['account.fiscalyear'].search(
                [('name', '=', tools.ustr(fiscal_year))])
            if fiscal_year_id:
                vals.update({'fiscalyear_id': fiscal_year_id.id})

            org = NonCostCtrl_Sheet.cell(row=2, column=2).value
            org_id =\
                self.env['res.org'].search(
                    ['|',
                     ('code', '=', tools.ustr(org)),
                     ('name_short', '=', tools.ustr(org))])
#             if org_id:
#                 vals.update({'org_id': org_id.id})

            section = NonCostCtrl_Sheet.cell(row=3, column=2).value
            section_id = self.env['res.section'].search(
                ['|',
                 ('code', '=', tools.ustr(section)),
                 ('name_short', '=', tools.ustr(org))])
            if section_id:
                vals.update({'section_id': section_id.id})
            export_date = NonCostCtrl_Sheet.cell(row=4, column=2).value
            if export_date:
                vals.update({'date': export_date})

            responsible_by = NonCostCtrl_Sheet.cell(row=5, column=2).value
            responsible_by_id = self.env['res.users'].search(
                [('name', '=', tools.ustr(responsible_by))])
            if section_id:
                vals.update({'creating_user_id': responsible_by_id.id})
            existing_lines = self.env['budget.plan.unit.line'].\
                search([('plan_id', '=', budget.id)])
            if existing_lines:
                existing_lines.unlink()
            line_row = 11
            lines_to_create = []
            for row in range(line_row, max_row):
                line_vals = {}
                line_vals.update({'section_id': section_id.id,
                                  'plan_id': budget.id,
                                  'org_id': org_id.id})
                ag_group = NonCostCtrl_Sheet.cell(row=row, column=1).value
                if not ag_group:
                    break
                ag_group_id = self.env['account.activity.group'].search(
                    [('name', '=', tools.ustr(ag_group))])
                if ag_group_id:
                    line_vals.update({'activity_group_id': ag_group_id.id})
                description = NonCostCtrl_Sheet.cell(row=row, column=2).value
                if description == '=FALSE()':
                    description = ''
                unit = NonCostCtrl_Sheet.cell(row=row, column=4).value or 0.0
                act_unitprice\
                    = NonCostCtrl_Sheet.cell(row=row, column=5).value or 0.0
                activity_unit =\
                    NonCostCtrl_Sheet.cell(row=row, column=6).value or 0.0
                line_vals.update({
                    'unit': unit,
                    'activity_unit_price': act_unitprice,
                    'activity_unit': activity_unit,
                    'description': description,
                })
                total_act_budget = unit * act_unitprice * activity_unit
                p = 1
                col = 9
                total_month_budget = 0.0
                while p != 13:
                    val = NonCostCtrl_Sheet.cell(row=row, column=col).value
                    if val and not isinstance(val, long):
                        raise UserError(
                            _('Please insert float value on\
                             row: %s - column: %s') % (row, col))
                    if val:
                        total_month_budget += val
                    line_vals.update({'m' + str(p): val})
                    if col == 20:
                        break
                    col += 1
                    p += 1
                lines_to_create.append(line_vals)
            for line in lines_to_create:
                self.env['budget.plan.unit.line'].create(line)
            budget.write(vals)
            attachement_id = self.env['ir.attachment'].create({
                'name': self.datas_fname,
                'datas': stream.getvalue().encode('base64'),
                'datas_fname': self.datas_fname,
                'res_model': 'budget.plan.unit',
                'res_id': budget.id,
            })
            self.env['budget.plan.history'].create({
                'user_id': self.env.user.id,
                'operation_date': fields.Datetime.now(),
                'operation_type': 'import',
                'plan_id': budget.id,
                'attachement_id': attachement_id.id
            })
            cost_control_data = self._update_cost_control_lines(workbook, budget)
        return {}

    @api.multi
    def import_budget(self):
        budget_ids = self.env.context.get('active_ids')
        result = self.update_budget_prepare(budget_ids, self.input_file)
        if result.get('messages'):
            msg = False
            for line in result['messages']:
                if not msg:
                    msg = line['message']
                else:
                    msg = msg + '\n' + line['message']
            if msg:
                raise UserError(_('%s') % (msg,))
