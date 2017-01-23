# -*- coding: utf-8 -*-
import base64
import cStringIO
from datetime import datetime

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from openpyxl.styles import PatternFill, Border, Side, Protection, Font
from openpyxl.utils import (_get_column_letter)

from openerp import models, fields, api, _
from openerp.exceptions import Warning as UserError

SHEET_FORMULAS = {}


class BudgetExportWizard(models.TransientModel):
    _name = 'unit.budget.plan.export'

    @api.model
    def _default_export_committed_budget(self):
        active_ids = self._context.get('active_ids', [])
        lines = len(self.env['budget.plan.unit.summary'].search(
            [('plan_id', 'in', active_ids)])._ids)
        if lines > 0:
            return False
        return True

    @api.model
    def _default_plan_template(self):
        template = self.env['ir.attachment'].\
            search([('budget_template', '=', True)])
        return template

    attachment_id = fields.Many2one(
        'ir.attachment',
        string='Template Plan',
        default=_default_plan_template,
    )
    editable_lines = fields.Integer(
        string='Additional Budget lines(Non JobOrder)',
        required=True,
        default=50,
    )
    export_committed_budget = fields.Boolean(
        string="Export Committed Budget?",
        default=_default_export_committed_budget,
    )

    @api.model
    def _compute_previous_year_amount(self, budget, budget_method):
        current_fy = budget.fiscalyear_id
        previous_fy = self.env['account.fiscalyear'].search(
            [('date_stop', '<', current_fy.date_start)],
            order='date_stop', limit=1)
        job_order_lines = {}
        non_job_order_lines = {}
        if not self.export_committed_budget or not previous_fy:
            return job_order_lines, non_job_order_lines
        # compute committed amount for previous year
        if previous_fy:
            report_domain = [('fiscalyear_id', '=', previous_fy.id),
                             ('section_id', '=', budget.section_id.id),
                             ('org_id', '=', budget.org_id.id),
                             ('division_id', '=', budget.division_id.id),
                             ('budget_method', '=', budget_method)]
            report_lines = self.env['budget.monitor.report'].\
                search(report_domain)
            for line in report_lines:
                total_commited_amt =\
                    line.amount_exp_commit +\
                    line.amount_po_commit +\
                    line.amount_pr_commit
                if total_commited_amt == 0.0:
                    continue
                if line.cost_control_id:
                    if line.cost_control_id.id not in job_order_lines:
                        job_order_lines[line.cost_control_id.id] = {}
                    if line.activity_group_id:
                        if line.activity_group_id.id not in \
                                job_order_lines[line.cost_control_id.id]:
                            job_order_lines[line.cost_control_id.id].\
                                update({line.activity_group_id.id:
                                        total_commited_amt})
                        else:
                            cc_id = line.cost_control_id.id
                            ag_id = line.activity_group_id.id
                            job_order_lines[cc_id][ag_id] += total_commited_amt
                    else:
                        if False not in \
                                job_order_lines[line.cost_control_id.id]:
                            job_order_lines[line.cost_control_id.id].\
                                update({False: total_commited_amt})
                        else:
                            cc_id = line.cost_control_id.id
                            job_order_lines[cc_id][False] += total_commited_amt
                else:
                    if line.activity_group_id:
                        if line.activity_group_id.id not in \
                                non_job_order_lines:
                            non_job_order_lines[line.activity_group_id.id] = \
                                total_commited_amt
                        else:
                            non_job_order_lines[line.activity_group_id.id] += \
                                total_commited_amt
                    else:
                        if False not in non_job_order_lines:
                            non_job_order_lines[False] = total_commited_amt
                        else:
                            non_job_order_lines[False] += total_commited_amt
        return job_order_lines, non_job_order_lines

    @api.model
    def _update_sheet_header(self, budget, Sheet):
        Sheet.cell(row=1, column=5, value=budget.id)
        org = budget.org_id.code or budget.org_id.name_short
        Sheet.cell(row=1, column=2, value=budget.fiscalyear_id.name)
        Sheet.cell(row=2, column=2, value=org)
        Sheet.cell(row=3, column=2, value=budget.section_id.code)
        Sheet.cell(row=4, column=2,
                   value=datetime.today().strftime('%d-%m-%Y'))
        Sheet.cell(row=5, column=2, value=self.env.user.name)

    @api.model
    def _update_sheet_lines(self, budget, budget_method, Sheet):
        ChargeTypeFormula = SHEET_FORMULAS.get('charge_type', False)
        ExpenseAGFormula = SHEET_FORMULAS.get('expense_ag_list', False)
        RevenueAGFormula = SHEET_FORMULAS.get('revenue_ag_list', False)
        JobOrderFormula = SHEET_FORMULAS.get('job_order_formula', False)

        Sheet.add_data_validation(ChargeTypeFormula)
        Sheet.add_data_validation(ExpenseAGFormula)
        Sheet.add_data_validation(RevenueAGFormula)
        Sheet.add_data_validation(JobOrderFormula)

        if budget_method == 'expense':
            lines = budget.plan_expense_line_ids
        else:
            lines = budget.plan_revenue_line_ids

        last_row = 10 + self.editable_lines
        row = 10
        job_order_lines, non_job_order_lines =\
            self._compute_previous_year_amount(budget, budget_method)
        if non_job_order_lines:
            for ag in non_job_order_lines:
                if ag:
                    ActivityGroup = self.env['account.activity.group']
                    ag_name = ActivityGroup.browse(ag).name
                    Sheet.cell(row=row, column=2).value = ag_name
                    Sheet.cell(row=row, column=10).value =\
                        non_job_order_lines[ag]
                else:
                    Sheet.cell(row=row, column=10).value =\
                        non_job_order_lines[ag]
                row += 1

        line_domain = [('id', 'not in', lines.ids),
                       ('plan_id', '=', budget.id),
                       ('budget_method', '=', budget_method),
                       ('cost_control_id', '!=', False)]
        joborderlines = self.env['budget.plan.unit.line'].search(line_domain)
        lines = lines + joborderlines

        for line in lines:
            ChargeTypeFormula.add(Sheet.cell(row=row, column=1))
            if budget_method == 'expense':
                ExpenseAGFormula.add(Sheet.cell(row=row, column=2))
            else:
                RevenueAGFormula.add(Sheet.cell(row=row, column=2))

            JobOrderFormula.add(Sheet.cell(row=row, column=3))
            charge_type =\
                line.charge_type == 'internal' and 'Internal' or 'External'

            Sheet.cell(row=row, column=1, value=charge_type)
            if line.activity_group_id:
                Sheet.cell(row=row, column=2,
                           value=line.activity_group_id.name)
            if line.cost_control_id:
                Sheet.cell(row=row, column=3, value=line.cost_control_id.name)
            if line.description:
                Sheet.cell(row=row, column=4, value=line.description)
            Sheet.cell(row=row, column=5, value='')
            Sheet.cell(row=row, column=6, value=line.unit)
            Sheet.cell(row=row, column=7, value=line.activity_unit_price)
            Sheet.cell(row=row, column=8, value=line.activity_unit)
#             Sheet.cell(row=row, column=9, value='')
#             Sheet.cell(row=row, column=10, value='')
            Sheet.cell(row=row, column=11, value=line.m1)
            Sheet.cell(row=row, column=12, value=line.m2)
            Sheet.cell(row=row, column=13, value=line.m3)
            Sheet.cell(row=row, column=14, value=line.m4)
            Sheet.cell(row=row, column=15, value=line.m5)
            Sheet.cell(row=row, column=16, value=line.m6)
            Sheet.cell(row=row, column=17, value=line.m7)
            Sheet.cell(row=row, column=18, value=line.m8)
            Sheet.cell(row=row, column=19, value=line.m9)
            Sheet.cell(row=row, column=20, value=line.m10)
            Sheet.cell(row=row, column=21, value=line.m11)
            Sheet.cell(row=row, column=22, value=line.m12)
            row += 1

        while row < last_row:
            ChargeTypeFormula.add(Sheet.cell(row=row, column=1))
            if budget_method == 'expense':
                ExpenseAGFormula.add(Sheet.cell(row=row, column=2))
            else:
                RevenueAGFormula.add(Sheet.cell(row=row, column=2))
            JobOrderFormula.add(Sheet.cell(row=row, column=3))
            row += 1

    @api.model
    def _format_sheet(self, Sheet):
        # Background Color
        CyanFill = PatternFill(
            start_color='E0FFFF',
            end_color='E0FFFF',
            fill_type='solid',
        )
        SkyBlueFill = PatternFill(
            start_color='71acd6',
            end_color='71acd6',
            fill_type='solid',
        )

        # Border Style
        thin = Side(
            border_style="thin",
            color="000000"
        )
        border = Border(
            top=thin,
            left=thin,
            right=thin,
            bottom=thin
        )

        # Font Style
        TahomaFont = Font(
            bold=False,
            name='Tahoma',
            size=10
        )
        TahomaRedFont = Font(
            bold=False,
            name='Tahoma',
            size=10,
            color='FF0000'
        )
        TahomaBoldFont = Font(
            bold=True,
            name='Tahoma',
            size=10
        )

        # Cell Type
        DecimalNumber = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1=0
        )

        num_format = '#,##0.00'
        ChargeTypeFormula = SHEET_FORMULAS.get('charge_type', False)
        ExpenseAGFormula = SHEET_FORMULAS.get('expense_ag_list', False)
        RevenueAGFormula = SHEET_FORMULAS.get('revenue_ag_list', False)
        JobOrderFormula = SHEET_FORMULAS.get('job_order_formula', False)

        # Adding drop-down formula to sheet
        Sheet.add_data_validation(ChargeTypeFormula)
        Sheet.add_data_validation(ExpenseAGFormula)
        Sheet.add_data_validation(RevenueAGFormula)
        Sheet.add_data_validation(JobOrderFormula)
        Sheet.add_data_validation(DecimalNumber)

        protection = Protection(locked=False)

        startrow = 10
        lastrow = self.editable_lines + startrow

        # Formating Lines
        for row in range(startrow, lastrow):
            for col in range(1, 26):
                Sheet.cell(row=row, column=col).fill = CyanFill
                Sheet.cell(row=row, column=col).border = border
                Sheet.cell(row=row, column=col).font = TahomaFont

                if col == 1:  # Charge Type Column
                    ChargeTypeFormula.add(Sheet.cell(row=row, column=col))
                    Sheet.cell(row=row, column=col).protection = protection
                elif col == 2:  # Activity Group Column
                    ExpenseAGFormula.add(Sheet.cell(row=row, column=col))
                    Sheet.cell(row=row, column=col).protection = protection
                elif col == 3:
                    JobOrderFormula.add(Sheet.cell(row=row, column=col))
                    Sheet.cell(row=row, column=col).protection = protection
                elif col == 9:  # Total Budget Column
                    col_F = _get_column_letter(6)
                    col_G = _get_column_letter(7)
                    col_H = _get_column_letter(8)
                    value = "=%s%s*$%s$%s*$%s$%s" % (col_F, row,
                                                     col_G, row,
                                                     col_H, row)
                    Sheet.cell(row=row, column=col).value = value
                    Sheet.cell(row=row, column=col).number_format = num_format
                elif col == 23:  # Total of phased entries Column
                    col_K = _get_column_letter(11)
                    col_V = _get_column_letter(22)
                    value = "=SUM(%s%s:$%s$%s)" % (col_K, row,
                                                   col_V, row)
                    Sheet.cell(row=row, column=col).value = value
                    Sheet.cell(row=row, column=col).number_format = num_format
                elif col == 24:  # Total minus phased total Column
                    col_I = _get_column_letter(9)
                    col_W = _get_column_letter(23)
                    value = "=IF(%s%s-%s%s<>0,%s%s-%s%s,0)" % (col_I, row,
                                                               col_W, row,
                                                               col_I, row,
                                                               col_W, row)
                    Sheet.cell(row=row, column=col).value = value
                    Sheet.cell(row=row, column=col).number_format = num_format
                elif col == 25:  # Error Column
                    col_X = _get_column_letter(24)
                    value = '=IF(ABS(%s%s)>0,"Error","")' % (col_X, row)
                    Sheet.cell(row=row, column=col).value = value
                    Sheet.cell(row=row, column=col).font = TahomaRedFont
                elif col not in (4, 5):
                    Sheet.cell(row=row, column=col).protection = protection
                    DecimalNumber.add(Sheet.cell(row=row, column=col))
                    Sheet.cell(row=row, column=col).number_format = num_format
                else:
                    Sheet.cell(row=row, column=col).protection = protection

        # Formating TOTAL line
        for col in range(1, 25):
            Sheet.cell(row=lastrow, column=col).fill = SkyBlueFill
            Sheet.cell(row=lastrow, column=col).border = border
            Sheet.cell(row=lastrow, column=col).font = TahomaFont

            if col == 1:
                Sheet.cell(row=lastrow, column=col).value = 'TOTAL'
                Sheet.cell(row=lastrow, column=col).font = TahomaBoldFont

            if col > 7:
                col_letter = _get_column_letter(col)
                value = "=SUM(%s%s:$%s$%s)" % (col_letter, startrow,
                                               col_letter, row)
                Sheet.cell(row=lastrow, column=col).value = value
                Sheet.cell(row=lastrow, column=col).number_format = num_format

    @api.model
    def _update_non_joborder_sheets(self, budget, workbook):
        # Update Expense sheet
        ExpenseSheet = workbook.get_sheet_by_name('Expense')
        ExpenseSheet.protection.sheet = True
        # set password to sheet
        ExpenseSheet.protection.set_password('pabi2')
        self._format_sheet(ExpenseSheet)
        self._update_sheet_header(budget, ExpenseSheet)
        self._update_sheet_lines(budget, 'expense', ExpenseSheet)

        # Update Revenue sheet
        RevenueSheet = workbook.get_sheet_by_name('Revenue')
        RevenueSheet.protection.sheet = True
        # set password to sheet
        RevenueSheet.protection.set_password('pabi2')
        self._format_sheet(RevenueSheet)
        self._update_sheet_header(budget, RevenueSheet)
        self._update_sheet_lines(budget, 'revenue', RevenueSheet)

    @api.model
    def _update_activity_group_sheet(self, workbook):
        try:
            AGObj = self.env['account.activity.group']
            AG_Sheet = workbook.get_sheet_by_name('Activity Group')
            AG_Sheet.protection.sheet = True
            AG_Sheet.protection.set_password('pabi2')
            expense_ags = AGObj.search([('budget_method', '=', 'expense')])
            # create lines in activity group data sheet
            expense_row = 2
            for ag in expense_ags:
                AG_Sheet.cell(row=expense_row, column=1, value=ag.name)
                AG_Sheet.cell(row=expense_row, column=2, value=ag.description)
                expense_row += 1
            formula1 = "{0}!$A$2:$A$%s" % (expense_row)
            ActGroupList = DataValidation(
                type="list",
                formula1=formula1.format(
                    quote_sheetname('Activity Group')
                )
            )
            SHEET_FORMULAS.update({'expense_ag_list': ActGroupList})
            revenue_ags = AGObj.search([('budget_method', '=', 'revenue')])
            # create lines in activity group data sheet
            revenue_row = expense_row + 1
            for ag in revenue_ags:
                AG_Sheet.cell(row=revenue_row, column=1, value=ag.name)
                AG_Sheet.cell(row=revenue_row, column=2, value=ag.description)
                revenue_row += 1
            revenue_formula = "{0}!$A$%s:$A$%s" % (expense_row+1, revenue_row)
            RevenueActGroupList = DataValidation(
                type="list",
                formula1=revenue_formula.format(
                    quote_sheetname('Activity Group')
                )
            )
            # Attach formulas to sheet to use further
            SHEET_FORMULAS.update({'revenue_ag_list': RevenueActGroupList})
        except:
            pass
        return True

    @api.model
    def _update_job_order_sheet(self, workbook):
        try:
            JobOrder_Sheet = workbook.get_sheet_by_name('JorOder')
            JobOrder_Sheet.protection.sheet = True
            JobOrder_Sheet.protection.set_password('pabi2')
            TahomaFont = Font(bold=False, name='Tahoma', size=10)
            joborders = self.env['cost.control'].search([])
            # Add cost control from odoo in cost control data sheet
            row = 2
            for joborder in joborders:
                JobOrder_Sheet.cell(
                    row=row,
                    column=1,
                    value=joborder.cost_control_type_id.name
                ).font = TahomaFont
                JobOrder_Sheet.cell(
                    row=row,
                    column=2,
                    value=joborder.code
                ).font = TahomaFont
                JobOrder_Sheet.cell(
                    row=row,
                    column=3,
                    value=joborder.name
                ).font = TahomaFont
                JobOrder_Sheet.cell(
                    row=row,
                    column=4,
                    value=joborder.name_short
                ).font = TahomaFont
                JobOrder_Sheet.cell(
                    row=row,
                    column=5,
                    value=''
                ).font = TahomaFont
                JobOrder_Sheet.cell(
                    row=row,
                    column=6,
                    value=''
                ).font = TahomaFont
                row += 1

            formula1 = "{0}!$C$2:$C$%s" % (row)
            JobOrdeListFormula = DataValidation(
                type="list",
                formula1=formula1.format(
                    quote_sheetname('JorOder')
                )
            )
            SHEET_FORMULAS.update({'job_order_formula': JobOrdeListFormula})
        except:
            pass
        return True

    @api.model
    def _get_filename(self, budget):
        template_file = self.attachment_id
        org = budget.org_id.code or budget.org_id.name_short
        filename = '%s-%s-%s-%s.xlsx' % (budget.fiscalyear_id.name,
                                         org,
                                         budget.section_id.code,
                                         template_file.name)
        return filename

    @api.multi
    def update_budget_xls(self, budget_ids, template_id=None):
        if not template_id:
            raise UserError(_('Please add .xlsx template.'))
        if not budget_ids:
            raise UserError(_('No budget to export.'))
        budgets = self.env['budget.plan.unit'].browse(budget_ids)

        for budget in budgets:
            exp_file = self.attachment_id.datas.decode('base64')
            stream = cStringIO.StringIO(exp_file)
            workbook = openpyxl.load_workbook(stream)

            ChargeType = DataValidation(
                type="list",
                formula1='"External,Internal"'
            )
            SHEET_FORMULAS.update({'charge_type': ChargeType})

            self._update_activity_group_sheet(workbook)
            self._update_job_order_sheet(workbook)
            self._update_non_joborder_sheets(budget, workbook)

            stream1 = cStringIO.StringIO()
            workbook.save(stream1)
            filename = self._get_filename(budget)
            self.env.cr.execute(""" DELETE FROM budget_xls_output """)
            attachement_id = self.env['ir.attachment'].create({
                'name': filename,
                'datas': stream1.getvalue().encode('base64'),
                'datas_fname': filename,
                'res_model': 'budget.plan.unit',
                'res_id': budget.id,
                'budget_plan_id': budget.id,
                'description': 'Export',
            })
            attach_id = self.env['budget.xls.output'].create({
                'name': filename,
                'xls_output': base64.encodestring(stream1.getvalue()),
            })
            self.env['budget.plan.history'].create({
                'user_id': self.env.user.id,
                'operation_date': fields.Datetime.now(),
                'operation_type': 'export',
                'plan_id': budget.id,
                'attachement_id': attachement_id.id
            })
            return attach_id

    @api.multi
    def export_budget(self):
        budget_ids = self.env.context.get('active_ids')
        attach_id = self.update_budget_xls(budget_ids, self.attachment_id.id)
        return {
            'context': self.env.context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'budget.xls.output',
            'res_id': attach_id.id,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }
