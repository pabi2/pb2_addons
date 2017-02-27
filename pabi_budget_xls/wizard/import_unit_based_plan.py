# -*- coding: utf-8 -*-
# © <YEAR(S)> <AUTHOR(S)>
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl.html).
import sys
import cStringIO
from datetime import datetime

import openpyxl

from openerp import tools
from openerp import models, fields, api, _
from openerp.exceptions import Warning as UserError


class BudgetImportWizard(models.TransientModel):
    _name = 'budget.import.wizard'

    input_file = fields.Binary('Template')
    datas_fname = fields.Char('Import File Name')

    @api.model
    def _create_joborder_lines(self, lines):
        groupby_job_order_dict = {}
        plan_id = False
        for line in lines:
            plan_id = line['plan_id']
            if line['cost_control_id'] not in groupby_job_order_dict:
                groupby_job_order_dict[line['cost_control_id']] = []
            groupby_job_order_dict[line['cost_control_id']].append(line)

        for joborder in groupby_job_order_dict:
            costcontrol_vals = {'plan_id': plan_id}
            costcontrol_vals.update({'cost_control_id': joborder})
            # get job order from each section
            existing_costcontrol_lines =\
                self.env['budget.plan.unit.cost.control'].search(
                    [('plan_id', '=', plan_id),
                     ('cost_control_id', '=', joborder)])
            # remove all job order line for plan
            if existing_costcontrol_lines:
                existing_costcontrol_lines.unlink()
            costcontrol_line = self.env['budget.plan.unit.cost.control'].\
                create(costcontrol_vals)
            calculate_vals = {}
            ids = []
            for line in groupby_job_order_dict[joborder]:
                # compute values for job order lines
                line_vals = {'cost_control_line_id': costcontrol_line.id}
                line_vals.update({
                    'charge_type': line['charge_type'],
                    'activity_group_id': line['activity_group_id'],
                    'name': line['description'],
                    'unit': line['unit'],
                    'activity_unit_price': line['activity_unit_price'],
                    'activity_unit': line['activity_unit'],
                    'm1': line['m1'],
                    'm2': line['m2'],
                    'm3': line['m3'],
                    'm4': line['m4'],
                    'm5': line['m5'],
                    'm6': line['m6'],
                    'm7': line['m7'],
                    'm8': line['m8'],
                    'm9': line['m9'],
                    'm10': line['m10'],
                    'm11': line['m11'],
                    'm12': line['m12'],
                })
                calculate_vals.update({
                    'description': line['description'],
                    'unit': line['unit'],
                    'activity_unit_price': line['activity_unit_price'],
                    'activity_unit': line['activity_unit'],
                    'budget_method': line['budget_method']
                })
                cc_line = self.env['budget.plan.unit.cost.control.line'].\
                    create(line_vals)
                if cc_line:
                    plan_line_vals = {}
                    plan_line_vals.update(line_vals)
                    plan_line_vals.update(calculate_vals)
                    create_vals = {
                        'breakdown_line_id': cc_line.id,
                        'plan_id': costcontrol_line.plan_id.id,
                        'fk_costcontrol_id': costcontrol_line.id,
                        'section_id': costcontrol_line.plan_id.section_id.id,
                        'cost_control_id': costcontrol_line.cost_control_id.id,
                    }
                    # compute values for plan job order lines and create lines
                    create_vals.update(plan_line_vals)
                    new_id = self.env['budget.plan.unit.line'].\
                        create(create_vals).id
                    ids.append(new_id)
            costcontrol_line.write({'detail_ids': [(6, 0, ids)]})
        return True

    @api.multi
    def update_budget_prepare(self, budget_ids, template=None):
        joborder_lines = []

        def _compute_line_vals(NonCostCtrl_Sheet, common_line_vals):
            line_row = 10
            lines_to_create = []
            max_row = NonCostCtrl_Sheet.max_row
            for row in range(line_row, max_row):
                line_vals = {}
                line_vals.update(common_line_vals)
                charge_type = NonCostCtrl_Sheet.cell(row=row, column=1).value
                if charge_type:
                    if str(charge_type) == 'External':
                        line_vals.update({'charge_type': 'external'})
                    else:
                        line_vals.update({'charge_type': 'internal'})
                ag_group = NonCostCtrl_Sheet.cell(row=row, column=2).value
                if not ag_group:
                    break
                ag_group_id = self.env['account.activity.group'].search(
                    [('name', '=', tools.ustr(ag_group))])
                if ag_group_id:
                    line_vals.update({'activity_group_id': ag_group_id.id})
                description = NonCostCtrl_Sheet.cell(row=row, column=4).value
                if description == '=FALSE()':
                    description = ''
                unit = NonCostCtrl_Sheet.cell(row=row, column=6).value or 0.0
                act_unitprice\
                    = NonCostCtrl_Sheet.cell(row=row, column=7).value or 0.0
                activity_unit =\
                    NonCostCtrl_Sheet.cell(row=row, column=8).value or 0.0
                line_vals.update({
                    'unit': unit,
                    'activity_unit_price': act_unitprice,
                    'activity_unit': activity_unit,
                    'description': description,
                })
                # total_act_budget = unit * act_unitprice * activity_unit
                p = 1
                col = 11
                total_month_budget = 0.0
                while p != 13:
                    val = NonCostCtrl_Sheet.cell(row=row, column=col).value
                    if val and not isinstance(val, long):
                        raise UserError(
                            _('Please insert float value on\
                             row: %s - column: %s') % (row, col))
                    if val:
                        total_month_budget += val
                    line_vals.update({'m' + str(p): val})
                    if col == 22:
                        break
                    col += 1
                    p += 1

                cost_control = NonCostCtrl_Sheet.cell(row=row, column=3).value
                if cost_control:
                    cost_control_id = self.env['cost.control'].search(
                        [('name', '=', tools.ustr(cost_control))])
                    if cost_control_id:
                        line_vals.update(
                            {'cost_control_id': cost_control_id.id})
                    joborder_lines.append(line_vals)
                else:
                    lines_to_create.append(line_vals)
            return lines_to_create

        if not template:
            raise UserError(_('Please add .xlsx template.'))
        if not budget_ids:
            raise UserError(_('No budget to import.'))

        imp_file = template.decode('base64')
        stream = cStringIO.StringIO(imp_file)

        try:
            workbook = openpyxl.load_workbook(stream)
        except IOError as e:
            raise UserError(_(e.strerror))
        except ValueError as e:
            raise UserError(_(e.strerror))
        except:
            e = sys.exc_info()[0]
            raise UserError(_('Wrong file format. Please enter .xlsx file.'))
        budgets = self.env['budget.plan.unit'].browse(budget_ids)
        for budget in budgets:
            # if we are trying to import after draft state then restrict him
            if budget.state != 'draft':
                raise UserError(
                    _('You can update budget plan only in draft state!'))

            Non_JobOrder_Expense = \
                workbook.get_sheet_by_name('Expense')
            vals = {}

            bg_id = Non_JobOrder_Expense.cell(row=1, column=5).value
            # if we trying to import sheet of other record then raise error
            if budget.id != bg_id:
                raise UserError(
                    _('Please import the correct file for this plan')
                )
            # get fiscal year from sheet
            fiscal_year = Non_JobOrder_Expense.cell(row=1, column=2).value
            fiscal_year_id = self.env['account.fiscalyear'].search(
                [('name', '=', tools.ustr(fiscal_year))])
            if fiscal_year_id:
                vals.update({'fiscalyear_id': fiscal_year_id.id})
            # get org from sheet
            org = Non_JobOrder_Expense.cell(row=2, column=2).value
            org_id =\
                self.env['res.org'].search(
                    ['|',
                     ('code', '=', tools.ustr(org)),
                     ('name_short', '=', tools.ustr(org))])
            # if org_id:
            #     vals.update({'org_id': org_id.id})
            # get section from sheet
            section = Non_JobOrder_Expense.cell(row=3, column=2).value
            section_id = self.env['res.section'].search(
                ['|',
                 ('code', '=', tools.ustr(section)),
                 ('name_short', '=', tools.ustr(org))])
            if section_id:
                vals.update({'section_id': section_id.id})
            export_date = Non_JobOrder_Expense.cell(row=4, column=2).value
            if export_date:
                export_date = datetime.strftime(
                    datetime.strptime(export_date, '%d-%m-%Y'), '%Y-%m-%d')
                vals.update({'date': export_date})
            responsible_by = Non_JobOrder_Expense.cell(row=5, column=2).value
            responsible_by_id = self.env['res.users'].search(
                [('name', '=', tools.ustr(responsible_by))])
            if section_id:
                vals.update({'creating_user_id': responsible_by_id.id})
            # remove plan lines from plan first
            existing_lines = self.env['budget.plan.unit.line'].\
                search([('plan_id', '=', budget.id)])
            if existing_lines:
                existing_lines.unlink()
            common_line_vals = {
                'section_id': section_id.id,
                'plan_id': budget.id,
                'org_id': org_id.id,
                'budget_method': 'expense',
            }
            # compute values for expense lines
            expense_lines_to_create = \
                _compute_line_vals(Non_JobOrder_Expense,
                                   common_line_vals=common_line_vals)
            common_line_vals['budget_method'] = 'revenue'
            Non_JobOrder_Revenue = \
                workbook.get_sheet_by_name('Revenue')
            # compute values for revenue lines
            revenue_lines_to_create = \
                _compute_line_vals(Non_JobOrder_Revenue,
                                   common_line_vals=common_line_vals)
            lines_to_create = expense_lines_to_create + revenue_lines_to_create
            self._create_joborder_lines(joborder_lines)
            # create all lines and assign to plan
            for line in lines_to_create:
                self.env['budget.plan.unit.line'].create(line)
            budget.write(vals)
            attachement_id = self.env['ir.attachment'].create({
                'name': self.datas_fname,
                'datas': stream.getvalue().encode('base64'),
                'datas_fname': self.datas_fname,
                'res_model': 'budget.plan.unit',
                'res_id': budget.id,
                'budget_plan_id': budget.id,
                'description': 'Import',
            })
            self.env['budget.plan.history'].create({
                'user_id': self.env.user.id,
                'operation_date': fields.Datetime.now(),
                'operation_type': 'import',
                'plan_id': budget.id,
                'attachement_id': attachement_id.id
            })
            # process for job order sheet
        return {}

    @api.multi
    def import_budget(self):
        budget_ids = self.env.context.get('active_ids')
        result = self.update_budget_prepare(budget_ids, self.input_file)
        if result.get('messages'):
            msg = False
            for line in result['messages']:
                if not msg:
                    msg = line['message']
                else:
                    msg = msg + '\n' + line['message']
            if msg:
                raise UserError(_('%s') % (msg,))
