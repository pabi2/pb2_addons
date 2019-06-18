# -*- coding: utf-8 -*-
import re
import os
import pandas as pd
import numpy as np
import xlrd
import csv
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Alignment, Font
from dateutil.parser import parse
from openpyxl.utils import get_column_interval
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl import load_workbook
import base64
import cStringIO
import time
from datetime import date, datetime as dt
from ast import literal_eval
from openerp.tools.float_utils import float_compare
from openerp import models, fields, api, _
from openerp.exceptions import except_orm, ValidationError
from openerp.tools.safe_eval import safe_eval
from openerp.addons.connector.queue.job import job, related_action
from openerp.addons.connector.session import ConnectorSession
from openerp.addons.connector.exception import FailedJobError
from openerp.addons.connector.exception import RetryableJobError

@job(default_channel='root.xlsx_report')
def action_done_async_process(session, model_name, res_id, lang=False):
    try:
        # Update context
        ctx = session.context.copy()
        if lang:
            ctx.update({'lang': lang})
        print '++++++++++++++++++++++++++++++++++', session, model_name, res_id
        out_file, out_name = session.pool[model_name].get_report(
            session.cr, session.uid, ctx)
        print '------------------------------out_file', out_file, out_name
        # Make attachment and link ot job queue
        job_uuid = session.context.get('job_uuid')
        job = session.env['queue.job'].search([('uuid', '=', job_uuid)],
                                              limit=1)
        # Get init time
        print '------------------------------job_uuid', job_uuid, job
        date_created = fields.Datetime.from_string(job.date_created)
        ts = fields.Datetime.context_timestamp(job, date_created)
        init_time = ts.strftime('%d/%m/%Y %H:%M:%S')
        # Create output report place holder
        desc = 'INIT: %s\n> UUID: %s' % (init_time, job_uuid)
        print '---------', out_name,'-', out_file, '-', job.id, '-', desc
        session.env['ir.attachment'].create({
            'name': out_name,
            'datas': out_file,
            'datas_fname': out_name,
            'res_model': 'queue.job',
            'res_id': job.id,
            'type': 'binary',
            'parent_id': session.env.ref('pabi_utils.dir_spool_report').id,
            'description': desc,
            'user_id': job.user_id.id,
        })
        # Result Description
        result = _('Successfully created excel report : %s') % out_name
        return result
    except Exception, e:
        raise FailedJobError(e)

    """try:
        res = session.pool[model_name].action_export(session.cr, session.uid, [res_id], session.context)
        return {'result': res}
    except Exception, e:
        raise RetryableJobError(e)"""


def adjust_cell_formula(value, k):
    """ Cell formula, i.e., if i=5, val=?(A11)+?(B12) -> val=A16+B17  """
    if isinstance(value, basestring):
        for i in range(value.count('?(')):
            if value and '?(' in value and ')' in value:
                i = value.index('?(')
                j = value.index(')', i)
                val = value[i + 2:j]
                col, row = split_row_col(val)
                new_val = '%s%s' % (col, row+k)
                value = value.replace('?(%s)' % val, new_val)
    return value


def get_field_aggregation(field):
    """ i..e, 'field@{sum/average/max/min}' """
    if field and '@{' in field and '}' in field:
        i = field.index('@{')
        j = field.index('}', i)
        cond = field[i + 2:j]
        try:
            if len(cond) > 0:
                return (field[:i], cond)
        except Exception:
            return (field.replace('@{%s}' % cond, ''), False)
    return (field, False)


def get_field_condition(field):
    """ i..e, 'field${value > 0 and value or False}' """
    if field and '${' in field and '}' in field:
        i = field.index('${')
        j = field.index('}', i)
        cond = field[i + 2:j]
        try:
            if len(cond) > 0:
                return (field.replace('${%s}' % cond, ''), cond)
        except Exception:
            return (field, False)
    return (field, False)


def get_field_format(field):
    """
        Available formats
        - font = bold, bold_red
        - fill = red, blue, yellow, green, grey
        - align = left, center, right
        - number = true, false

        i.e., 'field#{font=bold;fill=red;align=center;number_format=number}'
    """
    if field and '#{' in field and '}' in field:
        i = field.index('#{')
        j = field.index('}', i)
        cond = field[i + 2:j]
        try:
            if len(cond) > 0:
                return (field.replace('#{%s}' % cond, ''), cond)
        except Exception:
            return (field, False)
    return (field, False)


def fill_cell_format(field, field_format):
    avail_format = {
        'font': {
            'bold': Font(name="Arial", size=10, bold=True),
            'bold_red':
                Font(name="Arial", size=10, color=colors.RED, bold=True),
        },
        'fill': {
            'red': PatternFill("solid", fgColor="FF0000"),
            'grey': PatternFill("solid", fgColor="DDDDDD"),
            'yellow': PatternFill("solid", fgColor="FFFCB7"),
            'blue': PatternFill("solid", fgColor="9BF3FF"),
            'green': PatternFill("solid", fgColor="B0FF99"),
        },
        'align': {
            'left': Alignment(horizontal='left'),
            'center': Alignment(horizontal='center'),
            'right': Alignment(horizontal='right'),
        },
        'number_format': {
            'number': '#,##0.00',
            'date': 'dd/mm/yyyy',
            'datestamp': 'yyyy-mm-dd',
            'text': '@',
            'percent': '0.00%',
        },
    }
    formats = field_format.split(';')
    for f in formats:
        (key, value) = f.split('=')
        if key not in avail_format.keys():
            raise ValidationError(_('Invalid format type %s' % key))
        if value.lower() not in avail_format[key].keys():
            raise ValidationError(
                _('Invalid value %s for format type %s' % (value, key)))
        cell_format = avail_format[key][value]
        if key == 'font':
            field.font = cell_format
        if key == 'fill':
            field.fill = cell_format
        if key == 'align':
            field.alignment = cell_format
        if key == 'number_format':
            if value == 'text':
                try:
                    # In case value can't be encoded as utf, we do normal str()
                    field.value = field.value.encode('utf-8')
                except Exception:
                    field.value = str(field.value)
            field.number_format = cell_format


def get_line_max(line_field):
    """ i.e., line_field = line_ids[100], max = 100 else 0 """
    if line_field and '[' in line_field and ']' in line_field:
        i = line_field.index('[')
        j = line_field.index(']')
        max_str = line_field[i + 1:j]
        try:
            if len(max_str) > 0:
                return (line_field[:i], int(max_str))
            else:
                return (line_field, False)
        except Exception:
            return (line_field, False)
    return (line_field, False)


def get_groupby(line_field):
    """i.e., line_field = line_ids["a_id, b_id"], groupby = ["a_id", "b_id"]"""
    if line_field and '[' in line_field and ']' in line_field:
        i = line_field.index('[')
        j = line_field.index(']')
        groupby = literal_eval(line_field[i:j+1])
        return groupby
    return False


def split_row_col(pos):
    match = re.match(r"([a-z]+)([0-9]+)", pos, re.I)
    if not match:
        raise ValidationError(_('Position %s is not valid') % pos)
    col, row = match.groups()
    return col, int(row)


def get_sheet_by_name(book, name):
    """ Get sheet by name for openpyxl """
    i = 0
    for sheetname in book.sheetnames:
        if sheetname == name:
            return book.worksheets[i]
        i += 1
    raise ValidationError(_("'%s' sheet not found") % (name,))


def add_row_skips(vals):
    """ Add row skips for side column when match with \\skiprow """
    skips = {}
    # Find all skips requried for each rows
    for k, v in vals.iteritems():
        if not k or '\\skiprow' not in k:
            continue  # No skip for this field
        row = 0
        for value in v:
            num_skip = isinstance(value, basestring) and \
                value.count('\\skiprow') or 0
            if skips.get(row, 0) < num_skip:
                skips.update({row: num_skip})
            row += 1
    if not skips:
        return vals
    # Add skips for all other fields
    for k, v in vals.iteritems():
        row = 0
        for value in v:
            num_skip = isinstance(value, basestring) and \
                value.count('\\skiprow') or 0
            if skips.get(row, 0) > num_skip:
                value = str(value)
                for x in range(skips[row] - num_skip):
                    value += '\\skiprow'
                v[row] = value
            row += 1
    return vals


def isfloat(input):
    try:
        float(input)
        return True
    except ValueError:
        return False


def isinteger(input):
    try:
        int(input)
        return True
    except ValueError:
        return False


def isdatetime(input):
    try:
        if len(input) == 10:
            dt.strptime(input, '%Y-%m-%d')
        elif len(input) == 19:
            dt.strptime(input, '%Y-%m-%d %H:%M:%S')
        else:
            return False
        return True
    except ValueError:
        return False


def str_to_number(input):
    if isinstance(input, basestring):
        if ' ' not in input:
            if isdatetime(input):
                return parse(input)
            elif isinteger(input):
                if not (len(input) > 1 and input[:1] == '0'):
                    return int(input)
            elif isfloat(input):
                if not (input.find(".") > 2 and input[:1] == '0'):  # 00.123
                    return float(input)
    return input


def load_workbook_range(range_string, ws):
    """ Select worksheet range and return as pandas dataframe """
    col_start, col_end = re.findall("[A-Z]+", range_string)
    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])
    return pd.DataFrame(data_rows,
                        columns=get_column_interval(col_start, col_end))


def csv_from_excel(excel_content, delimiter, quote):
    decoded_data = base64.decodestring(excel_content)
    wb = xlrd.open_workbook(file_contents=decoded_data)
    sh = wb.sheet_by_index(0)
    content = cStringIO.StringIO()
    quoting = csv.QUOTE_ALL
    if not quote:
        quoting = csv.QUOTE_NONE
    wr = csv.writer(content, delimiter=delimiter, quoting=quoting)
    for rownum in xrange(sh.nrows):
        row_vals = map(lambda x: isinstance(x, basestring) and
                       x.encode('utf-8') or x,
                       sh.row_values(rownum))
        row = []
        for x in row_vals:
            if isinstance(x, basestring):
                x = x.strip()
            row.append(x)
        wr.writerow(row)
    # content.close()  # Set index to 0, and start reading
    content.seek(0)  # Set index to 0, and start reading
    out_file = base64.encodestring(content.read())
    return out_file


class ExportXlsxTemplate(models.TransientModel):
    """ This wizard is used with the template (ir.attachment) to export
    xlsx template filled with data form the active record """
    _name = 'export.xlsx.template'


    name = fields.Char(
        string='File Name',
        readonly=True,
        size=500,
    )
    data = fields.Binary(
        string='File',
        readonly=True,
    )
    template_id = fields.Many2one(
        'ir.attachment',
        string='Template',
        required=True,
        ondelete='cascade',
        domain="[('res_model', '=', res_model)]",
    )
    res_id = fields.Integer(
        string='Resource ID',
        readonly=True,
        required=True,
    )
    res_model = fields.Char(
        string='Resource Model',
        readonly=True,
        required=True,
        size=500,
    )
    async_process = fields.Boolean(
        string='Run task in background?',
        default=False,
    )
    uuid = fields.Char(
        string='UUID',
        readonly=True,
        size=500,
        help="Job queue unique identifier",
    )
    state = fields.Selection(
        [('choose', 'choose'),
         ('get', 'get')],
        default='choose',
    )


    @api.model
    def _get_template_fname(self):
        """ By default, get template_fname from context """
        template_fname = self._context.get('template_fname', False)
        return template_fname

    @api.model
    def default_get(self, fields):
        res_model = self._context.get('active_model', False)
        res_id = self._context.get('active_id', False)
        template_dom = [('res_model', '=', res_model),
                        ('parent_id', '!=', False)]
        template_fname = self._get_template_fname()
        if template_fname:  # Specific template
            template_dom.append(('datas_fname', '=', template_fname))
        templates = self.env['ir.attachment'].search(template_dom)
        if not templates:
            raise ValidationError(_('No template found!'))
        defaults = super(ExportXlsxTemplate, self).default_get(fields)
        for template in templates:
            if not template.datas:
                raise ValidationError(_('No file in %s') % (template.name,))
        defaults['template_id'] = len(templates) == 1 and templates.id or False
        defaults['res_id'] = res_id
        defaults['res_model'] = res_model
        return defaults

    @api.model
    def _get_line_vals(self, record, line_field, fields):
        """ Get values of this field from record set """
        line_field, max_row = get_line_max(line_field)
        lines = record[line_field]
        if max_row > 0 and len(lines) > max_row:
            raise Exception(
                _('Records in %s exceed max record allowed!') % line_field)
        vals = dict([(field, []) for field in fields])
        # Get field condition & aggre function
        field_cond_dict = {}
        aggre_func_dict = {}
        field_format_dict = {}
        pair_fields = []  # I.e., ('debit${value and . or .}@{sum}', 'debit')
        for field in fields:
            temp_field, eval_cond = get_field_condition(field)
            temp_field, field_format = get_field_format(temp_field)
            raw_field, aggre_func = get_field_aggregation(temp_field)
            # Dict of all special conditions
            field_cond_dict.update({field: eval_cond})
            aggre_func_dict.update({field: aggre_func})
            field_format_dict.update({field: field_format})
            # --
            pair_fields.append((field, raw_field))
        # --
        for line in lines:
            for field in pair_fields:  # (field, raw_field)
                value = self._get_field_data(field[1], line)
                # Case Eval
                eval_cond = field_cond_dict[field[0]]
                if eval_cond:  # Get eval_cond of a raw field
                    eval_context = {'float_compare': float_compare,
                                    'time': time,
                                    'datetime': dt,
                                    'date': date,
                                    'value': value,
                                    'object': line,
                                    'model': self.env[record._name],
                                    'env': self.env,
                                    'context': self._context,
                                    }
                    # value = str(eval(eval_cond, eval_context))
                    # Test removing str(), coz some case, need resulting number
                    value = eval(eval_cond, eval_context)
                # --
                vals[field[0]].append(value)
        return (vals, aggre_func_dict, field_format_dict)

    @api.model
    def _fill_workbook_data(self, workbook, record, data_dict):
        """ Fill data from record with format in data_dict to workbook """
        if not record or not data_dict:
            return
        try:
            # variable to store data range of each worksheet
            worksheet_range = {}
            for sheet_name in data_dict:
                ws = data_dict[sheet_name]
                st = False
                if isinstance(sheet_name, str):
                    st = get_sheet_by_name(workbook, sheet_name)
                elif isinstance(sheet_name, int):
                    st = workbook.worksheets[sheet_name - 1]
                if not st:
                    raise ValidationError(
                        _('Sheet %s not found!') % sheet_name)
                # ================ HEAD ================
                self._fill_head(ws, st, record)
                # ============= Line Items =============
                # Check for groupby directive
                groupbys = {key: ws[key] for key in
                            filter(lambda l: l[0:9] == '_GROUPBY_', ws.keys())}
                all_rc, max_row, tail_fields = self._fill_lines(ws, st, record,
                                                                groupbys)
                # ================ TAIL ================
                self._fill_tail(ws, st, record, tail_fields)

                # prepare worksheet data range, to be used in BI funtions
                if all_rc:
                    begin_rc = min(all_rc)
                    col, row = split_row_col(
                        max(sorted(all_rc, reverse=True), key=len))
                    end_rc = '%s%s' % (col, max_row)
                    worksheet_range[sheet_name] = '%s:%s' % (begin_rc, end_rc)

            # ================ BI Function ================
            self._fill_bi(workbook, data_dict, worksheet_range)

        except KeyError, e:
            raise except_orm(_('Key Error!'), e)
        except IllegalCharacterError, e:
            raise except_orm(
                _('IllegalCharacterError!\n'
                  'Some exporting data may contain special character'), e)
        except Exception, e:
            raise except_orm(_('Error filling data into excel sheets!'), e)

    @api.model
    def _get_field_data(self, _field, _line):
        """ Get field data, and convert data type if needed """
        if not _field:
            return None
        line_copy = _line
        for f in _field.split('.'):
            data_type = line_copy._fields[f].type
            line_copy = line_copy[f]
            if data_type == 'date':
                if line_copy:
                    line_copy = dt.strptime(line_copy, '%Y-%m-%d')
            elif data_type == 'datetime':
                if line_copy:
                    line_copy = dt.strptime(line_copy, '%Y-%m-%d %H:%M:%S')
        if isinstance(line_copy, basestring):
            line_copy = line_copy.encode('utf-8')
        return line_copy

    @api.model
    def _fill_head(self, ws, st, record):
        for rc, field in ws.get('_HEAD_', {}).iteritems():
            tmp_field, eval_cond = get_field_condition(field)
            tmp_field, field_format = get_field_format(tmp_field)
            value = tmp_field and self._get_field_data(tmp_field, record)
            # Case Eval
            if eval_cond:  # Get eval_cond of a raw field
                eval_context = {'float_compare': float_compare,
                                'time': time,
                                'datetime': dt,
                                'date': date,
                                'value': value,
                                'object': record,
                                'model': self.env[record._name],
                                'env': self.env,
                                'context': self._context,
                                }
                # str() throw cordinal not in range error
                value = eval(eval_cond, eval_context)
                # value = str(eval(eval_cond, eval_context))
            if value is not None:
                st[rc] = value
            if field_format:
                fill_cell_format(st[rc], field_format)

    @api.model
    def _fill_lines(self, ws, st, record, groupbys):
        all_rc = []
        max_row = 0

        line_fields = ws.keys()
        for x in ('_HEAD_', '_TAIL_', '_GROUPBY_', '_BI_'):
            for line_field in line_fields:
                if x in line_field:
                    line_fields.remove(line_field)
        tail_fields = {}  # Keep tail cell, to be used in _TAIL_
        for line_field in line_fields:

            subtotals = {'rows': [], 'subtotals': {},
                         'grandtotals': {}, 'formats': {}}
            # ====== GROUP BY =========
            groupby_keys = filter(lambda l: '_GROUPBY_%s' % line_field in l,
                                  groupbys.keys())
            if groupby_keys:
                groupby = get_groupby(groupby_keys[0])
                groupby_dict = groupbys[groupby_keys[0]]
                # If value in groupby changes, mark the row index
                i = 0
                old_val = []
                for line in record[line_field]:
                    val = []
                    for key in groupby:
                        val.append(line[key])
                    if i > 0 and val != old_val:
                        subtotals['rows'].append(i)
                    old_val = val
                    i += 1
                subtotals['rows'].append(i)
                # For the aggregrate column
                for cell in groupby_dict.keys():
                    col, row = split_row_col(cell)
                    first_row = row
                    cell_format = groupby_dict[cell]
                    _, grp_func = get_field_aggregation(cell_format)
                    _, grp_format = get_field_format(cell_format)

                    cell_field = ws[line_field][cell]
                    subtotals['subtotals'].update({cell_field: []})
                    subtotals['formats'].update({cell_field: []})
                    grandtotals = []
                    if subtotals['rows']:
                        subtotals['formats'][cell_field] = grp_format
                    for i in subtotals['rows']:
                        from_cell = '%s%s' % (col, row)
                        to_cell = '%s%s' % (col, first_row+i-1)
                        range_cell = '%s:%s' % (from_cell, to_cell)
                        subtotals['subtotals'][cell_field].append(
                            grp_func and
                            '=%s(%s)' % (grp_func, range_cell) or
                            '')
                        grandtotals.append(range_cell)
                        first_row += 1
                        row = first_row + i
                    if grandtotals:
                        subtotals['grandtotals'][cell_field] = \
                            '=%s(%s)' % (grp_func, ','.join(grandtotals))
                # --

            sb_rows = [i + v for i, v in enumerate(subtotals.get('rows', []))]
            sb_subtotals = subtotals.get('subtotals', {})
            sb_grandtotals = subtotals.get('grandtotals', {})
            sb_formats = subtotals.get('formats', {})

            fields = ws.get(line_field, {}).values()
            all_rc += ws.get(line_field, {}).keys()
            (vals, func, field_format) = \
                self._get_line_vals(record, line_field, fields)
            # value with '\\skiprow' signify line skipping
            vals = add_row_skips(vals)

            for rc, field in ws.get(line_field, {}).iteritems():
                tail_fields[rc] = False
                col, row = split_row_col(rc)  # starting point
                i = 0
                new_row = 0
                new_rc = rc
                for val in vals[field]:
                    row_vals = isinstance(val, basestring) and \
                        val.split('\\skiprow') or [val]
                    for row_val in row_vals:
                        new_row = row + i
                        new_rc = '%s%s' % (col, new_row)
                        row_val = adjust_cell_formula(row_val, i)
                        if row_val not in ('None', None):
                            st[new_rc] = str_to_number(row_val)
                        if field_format.get(field, False):
                            fill_cell_format(st[new_rc],
                                             field_format[field])
                        # ====== GROUP BY =========
                        # Sub Total
                        for j in sb_rows:
                            if i == j-1:
                                new_row = row + i
                                if sb_subtotals.get(field, False):
                                    new_rc = '%s%s' % (col, new_row + 1)
                                    row_val = sb_subtotals[field][0]
                                    st[new_rc] = str_to_number(row_val)
                                    sb_subtotals[field].pop(0)
                                if sb_formats.get(field, False):
                                    new_rc = '%s%s' % (col, new_row + 1)
                                    grp_format = sb_formats[field]
                                    if grp_format:
                                        fill_cell_format(st[new_rc],
                                                         grp_format)
                                i += 1

                        # ---------------------------
                        if new_row > max_row:
                            max_row = new_row
                        i += 1

                # Grand Total
                if sb_grandtotals.get(field, False):
                    new_rc = '%s%s' % (col, new_row + 2)
                    row_val = sb_grandtotals[field]
                    st[new_rc] = str_to_number(row_val)
                if sb_formats.get(field, False):
                    new_rc = '%s%s' % (col, new_row + 2)
                    grp_format = sb_formats[field]
                    if grp_format:
                        fill_cell_format(st[new_rc],
                                         grp_format)

                # Add footer line if at least one field have func
                f = func.get(field, False)
                if f:
                    new_row += 1
                    f_rc = '%s%s' % (col, new_row)
                    st[f_rc] = '=%s(%s:%s)' % (f, rc, new_rc)

                tail_fields[rc] = new_rc   # Last row field

        return all_rc, max_row, tail_fields

    @api.model
    def _fill_tail(self, ws, st, record, tail_fields):
        # Get the max last rc's row
        last_row = 0
        for to_rc in tail_fields.values():
            _, row = split_row_col(to_rc)
            last_row = row > last_row and row or last_row
        # Similar to header, except it will set cell after last row
        # Get all tails, i.e., _TAIL_0, _TAIL_1 order by number
        tails = filter(lambda l: l[0:6] == '_TAIL_', ws.keys())
        tail_dicts = {key: ws[key] for key in tails}
        for tail_key, tail_dict in tail_dicts.iteritems():
            row_skip = tail_key[6:] != '' and int(tail_key[6:]) or 0
            # For each _TAIL_ and row skipper 0, 1, 2, ...
            for rc, field in tail_dict.iteritems():
                tmp_field, eval_cond = get_field_condition(field)
                tmp_field, field_format = get_field_format(tmp_field)
                tmp_field, func = get_field_aggregation(tmp_field)
                value = tmp_field and self._get_field_data(tmp_field, record)
                # Case Eval
                if eval_cond:  # Get eval_cond of a raw field
                    eval_context = {'float_compare': float_compare,
                                    'time': time,
                                    'datetime': dt,
                                    'date': date,
                                    'value': value,
                                    'object': record,
                                    'model': self.env[record._name],
                                    'env': self.env,
                                    'context': self._context,
                                    }
                    # str() throw cordinal not in range error
                    value = eval(eval_cond, eval_context)
                    # value = str(eval(eval_cond, eval_context))
                # If no rc in tail_fields, use the max last row
                last_rc = False
                tail_rc = False
                if rc in tail_fields.keys():
                    last_rc = tail_fields[rc]  # Last row of rc column
                    col, row = split_row_col(last_rc)
                    tail_rc = '%s%s' % (col, row + row_skip + 1)
                else:
                    col, _ = split_row_col(rc)
                    last_rc = '%s%s' % (col, last_row)
                    tail_rc = '%s%s' % (col, last_row + row_skip + 1)
                if value and value is not None:
                    st[tail_rc] = value
                if func:
                    st[tail_rc] = '=%s(%s:%s)' % (func, rc, last_rc)
                if field_format:
                    fill_cell_format(st[tail_rc], field_format)

    @api.model
    def _fill_bi(self, workbook, data_dict, worksheet_range):
        for sheet_name in data_dict:
            worksheet = data_dict[sheet_name]
            if isinstance(sheet_name, str):
                st = get_sheet_by_name(workbook, sheet_name)
            elif isinstance(sheet_name, int):
                st = workbook.worksheets[sheet_name - 1]
            if not st:
                raise ValidationError(
                    _('Sheet %s not found!') % sheet_name)
            if not worksheet.get('_BI_', False):
                continue
            for rc, bi_dict in worksheet.get('_BI_', {}).iteritems():
                req_field = ['df', 'oper_code']
                key_field = bi_dict.keys()
                if set(req_field) != set(key_field):
                    raise ValidationError(
                        _('_BI_ requires \n'
                          ' - df: initial DataFrame from worksheet\n'
                          ' - oper_code: pandas operation code'))
                # Get dataframe
                src_df = bi_dict['df']
                src_st = get_sheet_by_name(workbook, src_df)
                df = load_workbook_range(worksheet_range[src_df], src_st)
                eval_context = {'df': df, 'pd': pd, 'np': np}
                # Get DF using safe_eval method
                df = safe_eval(bi_dict['oper_code'], eval_context,
                               mode="exec", nocopy=True)
                if 'result' in eval_context:  # use result=...
                    df = eval_context['result']
                if df is None:
                    df = eval(bi_dict['oper_code'], eval_context)
                if df.empty:
                    continue
                df = df.reset_index()
                rows = dataframe_to_rows(df, index=False, header=False)
                # Get init cell index
                xy = coordinate_from_string(rc)
                c = column_index_from_string(xy[0])
                r = xy[1]
                for r_idx, row in enumerate(rows, r):
                    for c_idx, value in enumerate(row, c):
                        st.cell(row=r_idx, column=c_idx, value=value)

    @api.model
    def _export_template(self, template, res_model, res_id,
                         to_csv=False, csv_delimiter=',',
                         csv_extension='csv', csv_quote=True):
        data_dict = literal_eval(template.description.strip())
        export_dict = data_dict.get('__EXPORT__', False)
        out_name = template.name
        if not export_dict:  # If there is not __EXPORT__ formula, just export
            out_name = template.datas_fname
            out_file = template.datas
            return (out_file, out_name)
        # Prepare temp file (from now, only xlsx file works for openpyxl)
        decoded_data = base64.decodestring(template.datas)
        ConfParam = self.env['ir.config_parameter']
        ptemp = ConfParam.get_param('path_temp_file') or '/temp'
        stamp = dt.utcnow().strftime('%H%M%S%f')[:-3]
        ftemp = '%s/temp%s.xlsx' % (ptemp, stamp)
        f = open(ftemp, 'w')
        f.write(decoded_data)
        f.seek(0)
        f.close()
        # Workbook created, temp fie removed
        wb = load_workbook(ftemp)
        os.remove(ftemp)
        # ============= Start working with workbook =============
        record = res_model and self.env[res_model].browse(res_id) or False
        self._fill_workbook_data(wb, record, export_dict)
        # =======================================================
        # Return file as .xlsx
        content = cStringIO.StringIO()
        wb.save(content)
        content.seek(0)  # Set index to 0, and start reading
        out_file = base64.encodestring(content.read())
        if record and 'name' in record and record.name:
            out_name = record.name.replace(' ', '').replace('/', '')
        else:
            fname = out_name.replace(' ', '').replace('/', '')
            ts = fields.Datetime.context_timestamp(self, dt.now())
            out_name = '%s_%s' % (fname, ts.strftime('%Y%m%d_%H%M%S'))
        if not out_name or len(out_name) == 0:
            out_name = 'noname'
        out_ext = 'xlsx'
        # CSV (convert only 1st sheet)
        if to_csv:
            delimiter = csv_delimiter.encode("utf-8")
            out_file = csv_from_excel(out_file, delimiter, csv_quote)
            out_ext = csv_extension
        return (out_file, '%s.%s' % (out_name, out_ext))



    @api.multi
    def get_report(self):
        self.ensure_one()
        Export = self.env['export.xlsx.template']
        Attachment = self.env['ir.attachment']
        template = []
        # By default, use template by model
        if self.template_id:
            template = self.template_id
        else:
            template = Attachment.search([('res_model', '=', self._name)])
        if len(template) != 1:
            raise ValidationError(
                _('No one template selected for "%s"') % self._name)
        return Export._export_template(
            template, self._name, self.id,
            to_csv=self.to_csv,
            csv_delimiter=self.csv_delimiter,
            csv_extension=self.csv_extension,
            csv_quote=self.csv_quote)

    @api.multi
    def act_getfile(self):
        self.ensure_one()
        if self.async_process == True:
            #if self._context.get('job_uuid', False):  # Called from @job
            #    return self.act_getfile()
            Job = self.env['queue.job']
            session = ConnectorSession(self._cr, self._uid, self._context)
            description = 'Excel Report - %s' % (self.res_model or self.name)
            uuid = action_done_async_process.delay(session, self._name, self.id, description=description, lang=session.context.get('lang', False))
            job = Job.search([('uuid', '=', uuid)], limit=1)
            # Process Name
            job.process_id = self.env.ref('pabi_utils.xlsx_report')
            self.write({'state': 'get', 'uuid': uuid})
        else:
            out_file, out_name = self._export_template(self.template_id, self.res_model, self.res_id)
            self.write({'state': 'get', 'data': out_file, 'name': out_name})
        print '--++--++--++--++---++----++----++----++--++---++---++--++--', self.id, self._name
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name, #'export.xlsx.template',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
